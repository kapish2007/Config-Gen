from jinja2 import Template
import csv
import ipaddress
from openpyxl import load_workbook
import os

def read_csv(file_path):
    data = []
    with open(file_path, newline='') as csvfile:
        csvreader = csv.DictReader(csvfile)
        for row in csvreader:
            if 'dhcp_helper' in row:
                row['dhcp_helper'] = row['dhcp_helper'].split(',')
            data.append(row)
    return data

def calculate_ip_addresses(vlan_network):
    network = ipaddress.IPv4Network(vlan_network)
    hsrp_vip = network.network_address + 1
    if network.prefixlen == 31:
        vlan_ip_pri = network.network_address
        vlan_ip_sec = network.network_address + 1
    else:
        vlan_ip_pri = network.network_address + 2
        vlan_ip_sec = network.network_address + 3
    mask = network.netmask
    net_add = network.network_address
    wildcard = network.hostmask
    return hsrp_vip, vlan_ip_pri, vlan_ip_sec, mask, net_add, wildcard

def read_cable_matrix(file_path):
    cable_matrix = []
    wb = load_workbook(filename=file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        source_device, source_floor, source_port, destination_device, destination_floor, destination_port = row
        cable_matrix.append({
            'source_device': source_device,
            'source_floor': source_floor,
            'source_port': source_port,
            'destination_device': destination_device,
            'destination_floor': destination_floor,
            'destination_port': destination_port,
        })
    return cable_matrix

def generate_core_switch_configs(template_file, switch_data, vlan_data, cable_matrix):
    with (open(template_file, 'r') as f):
        template_content = f.read()
        jinja_template = Template(template_content)
        config_files = []
        for switch_values in switch_data:
            cable_data = []
            for cable in cable_matrix:
                if cable['source_device'] == switch_values['hostname']:
                    cable_data.append(cable)
            for vlan in vlan_data:
                (hsrp_vip, vlan_ip_pri, vlan_ip_sec, mask, net_add, wildcard) = calculate_ip_addresses(vlan['network'])
                vlan['hsrp_vip'] = hsrp_vip
                vlan['vlan_ip_pri'] = vlan_ip_pri
                vlan['vlan_ip_sec'] = vlan_ip_sec
                vlan['mask'] = mask
                vlan['net_add'] = net_add
                vlan['wildcard'] = wildcard
            populated_template = jinja_template.render(switch=switch_values, vlans=vlan_data, cables=cable_data)
            config_files.append((switch_values['hostname'], populated_template))
        return config_files

def generate_access_switch_configs(template_file, switch_data, vlan_data):
    with (open(template_file, 'r') as f):
        template_content = f.read()
        jinja_template = Template(template_content)
        config_files = []
        for switch_values in switch_data:
            populated_template = jinja_template.render(switch=switch_values, vlans=vlan_data)
            config_files.append((switch_values['hostname'], populated_template))
        return config_files
def save_configs(path, configs):
    for hostname, config in configs:
        output_file = os.path.join(path, f'{hostname}.txt')
        with open(output_file, 'w') as f:
            f.write(config)
            print(f"Configuration for '{hostname}' generated & saved to {output_file}")

def main():
    print("\nWelcome to v1.0 of this script!")
    print("\n*** DISCLAIMER : This script caters only to ASPAC region!!! ***")
    dir_path = input("\nInput directory path for all files:")
    #example - /Users/kapish/Documents/Scripts/
    while True:
        print("\nMain Menu:")
        print("1. Generate KV Access (Layer 2) Switch Configuration")
        print("2. Generate KV Core (Layer 3) Switch Configuration")
        print("3. Generate JNJ Access (Layer 2) Switch Configuration")
        print("4. Generate JNJ Core (Layer 3) Switch Configuration")
        print("5. Exit")
        choice = input("\nSelect an option: ")

        if choice == '1':
            switch_csv_file_path = os.path.join(dir_path, 'access_switch_config.csv')
            switch_data = read_csv(switch_csv_file_path)

            vlan_csv_file_path = os.path.join(dir_path, 'vlan_info.csv')
            vlan_data = read_csv(vlan_csv_file_path)

            template_file = os.path.join(dir_path, 'access-switch.j2')

            access_switch_config = generate_access_switch_configs(template_file, switch_data, vlan_data)
            save_configs(dir_path, access_switch_config)
            print("\nLayer 2 Switch configurations are completed. Please review & deploy with care. Good Bye!")
            break

        elif choice == '2':
            switch_csv_file_path = os.path.join(dir_path, 'core_switch_config.csv')
            switch_data = read_csv(switch_csv_file_path)

            vlan_csv_file_path = os.path.join(dir_path, 'vlan_info.csv')
            vlan_data = read_csv(vlan_csv_file_path)

            cable_matrix_file_path = os.path.join(dir_path, 'cable_matrix.xlsx')
            cable_matrix = read_cable_matrix(cable_matrix_file_path)

            template_file = os.path.join(dir_path, 'core-switch.j2')

            core_switch_config = generate_core_switch_configs(template_file, switch_data, vlan_data, cable_matrix)
            save_configs(dir_path, core_switch_config)
            print("\nLayer 3 Switch configurations are completed. Please review & deploy with care. Good Bye!")
            break

        elif choice == '3':
            switch_csv_file_path = os.path.join(dir_path, 'access_switch_config.csv')
            switch_data = read_csv(switch_csv_file_path)

            vlan_csv_file_path = os.path.join(dir_path, 'vlan_info.csv')
            vlan_data = read_csv(vlan_csv_file_path)

            template_file = os.path.join(dir_path, 'jnj-access-switch.j2')

            access_switch_config = generate_access_switch_configs(template_file, switch_data, vlan_data)
            save_configs(dir_path, access_switch_config)
            print("\nLayer 2 Switch configurations are completed. Please review & deploy with care. Good Bye!")
            break

        elif choice == '4':
            switch_csv_file_path = os.path.join(dir_path, 'core_switch_config.csv')
            switch_data = read_csv(switch_csv_file_path)

            vlan_csv_file_path = os.path.join(dir_path, 'vlan_info.csv')
            vlan_data = read_csv(vlan_csv_file_path)

            cable_matrix_file_path = os.path.join(dir_path, 'cable_matrix.xlsx')
            cable_matrix = read_cable_matrix(cable_matrix_file_path)

            template_file = os.path.join(dir_path, 'jnj-core-switch.j2')

            core_switch_config = generate_core_switch_configs(template_file, switch_data, vlan_data, cable_matrix)
            save_configs(dir_path, core_switch_config)
            print("\nLayer 3 Switch configurations are completed. Please review & deploy with care. Good Bye!")
            break

        elif choice == '5':
            print("\nExiting, Good Bye!")
            break

        else:
            print("Invalid choice. Please select a valid option.")

if __name__ == '__main__':
    main()
